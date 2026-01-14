import sys
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADERS: List[str] = [
    "Date",
    "Issues Triaged",
    "Issues Resolved",
    "PRs Created",
    "PRs Merged",
    "Commits",
    "Open Issues",
    "Open PRs",
    "ADO Tests",
    "Release",
    "Notes",
]


def die(msg: str, code: int = 1) -> None:
    print(f"[init_tracker] ERROR: {msg}", file=sys.stderr)
    sys.exit(code)


def load_dotenv(dotenv_path: Path) -> Dict[str, str]:
    if not dotenv_path.is_file():
        die(f".env file not found at: {dotenv_path}")

    data: Dict[str, str] = {}
    for raw in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip()

        # Strip matching quotes
        if len(v) >= 2 and v[0] == v[-1] and v[0] in ("'", '"'):
            v = v[1:-1].strip()

        data[k] = v
    return data


def require(cfg: Dict[str, str], key: str) -> str:
    v = cfg.get(key, "").strip()
    if not v:
        die(f"Missing required config in .env: {key}")
    return v


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="305496")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")

    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = fill
        c.font = font
        c.alignment = align

    ws.freeze_panes = "A2"

    widths = [12, 14, 15, 12, 12, 10, 12, 10, 14, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    cfg = load_dotenv(script_dir / ".env")

    out = require(cfg, "TRACKER_OUT")
    sheet_name = require(cfg, "TRACKER_SHEET")

    owner = require(cfg, "GITHUB_OWNER")
    repo = require(cfg, "GITHUB_REPO")
    username = require(cfg, "GITHUB_USERNAME")
    timezone = require(cfg, "TRACKER_TIMEZONE")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    style_header(ws)

    config_ws = wb.create_sheet("Config")
    config_ws["A1"] = "GitHub Owner"
    config_ws["A2"] = "GitHub Repo"
    config_ws["A3"] = "GitHub Username"
    config_ws["A4"] = "Timezone"
    config_ws["A5"] = "Last Updated (UTC)"
    for cell in ["A1", "A2", "A3", "A4", "A5"]:
        config_ws[cell].font = Font(bold=True)

    config_ws["B1"] = owner
    config_ws["B2"] = repo
    config_ws["B3"] = username
    config_ws["B4"] = timezone

    config_ws.column_dimensions["A"].width = 22
    config_ws.column_dimensions["B"].width = 28

    wb.save(out)
    print(f"[init_tracker] Created workbook: {out}")


if __name__ == "__main__":
    main()
