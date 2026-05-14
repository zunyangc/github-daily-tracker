"""
GitHub -> Excel daily tracker updater (with progress logging)

Purpose
-------
Updates (or creates) a single row in your Excel tracker for a given day, filling
in GitHub-derived metrics for *your personal contributions*.

What this script updates in Excel (auto-populated)
--------------------------------------------------
- Issues Triaged:
    Unique issues in the target repo where you commented on that day.
    (Uses Issue Comments API: GET /repos/.../issues/comments, filtered by user + date)
- Issues Resolved:
    Issues you closed on that day in the target repo.
    (Uses Search API + Issue Events API to verify you were the closer)
- PRs Created:
    PRs you opened that day.
    (Uses Search API: author:USER created:DAY)
- PRs Merged:
    PRs you authored that were merged that day.
    (Uses Search API: author:USER merged:DAY)
- Commits:
    Commits you authored in the target repo that day.
    (Uses Commits API: GET /repos/.../commits?author=USER&since=&until=)
- Open Issues / Open PRs (as-of that day):
    Snapshot count approximated by summing two queries:
      (1) created<=day AND still open  (2) created<=day AND closed after day

Column layout (data sheet)
--------------------------
- A Date            (written + format normalized every run)
- B Issues Triaged  (written)
- C Issues Resolved (written)
- D PRs Created     (written)
- E PRs Merged      (written)
- F Commits         (written)
- G Open Issues     (written)
- H Open PRs        (written)
- I Release         (manual — NEVER touched)
- J Notes           (manual — NEVER touched)

Defaults
--------
- If no date argument is provided, defaults to *today* (local machine date).

Environment requirements
------------------------
- Set GITHUB_TOKEN (GitHub PAT) in env or in a .env file loaded by your runner.
"""

import os
import sys
import time
import datetime as dt
import zipfile
from typing import Tuple, Dict, Any, List
from pathlib import Path
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import requests
from openpyxl import load_workbook

DEFAULT_API = "https://api.github.com"
DEFAULT_DATE_FORMAT = "dd/mm/yyyy"
DEFAULT_MAX_PAGES = 10
DEFAULT_TIMEZONE = "UTC"


def resolve_tz(tz_name: str) -> ZoneInfo:
    """Resolve a timezone name to a ZoneInfo, dying with a helpful message on failure."""
    try:
        return ZoneInfo(tz_name)
    except ZoneInfoNotFoundError:
        die(f"Unknown TRACKER_TIMEZONE '{tz_name}'. Use an IANA name like 'Asia/Kuala_Lumpur' or 'UTC'.")


def day_bounds_utc(day: dt.date, tz: ZoneInfo) -> Tuple[dt.datetime, dt.datetime]:
    """
    Return [start_utc, end_utc) covering the local-day `day` in timezone `tz`.

    The script treats a "day" as the user's local day. All API queries are
    rewritten to use this UTC window so that contributions near local midnight
    land on the expected tracker date.
    """
    start_local = dt.datetime.combine(day, dt.time.min, tzinfo=tz)
    end_local = start_local + dt.timedelta(days=1)
    return (
        start_local.astimezone(dt.timezone.utc).replace(tzinfo=None),
        end_local.astimezone(dt.timezone.utc).replace(tzinfo=None),
    )


def fmt_iso_z(t: dt.datetime) -> str:
    """Format a naive UTC datetime as ISO-8601 with trailing Z (GitHub-friendly)."""
    return t.strftime("%Y-%m-%dT%H:%M:%SZ")


def load_dotenv(dotenv_path: str = ".env") -> Dict[str, str]:
    """
    Minimal .env loader.
    Reads KEY=VALUE pairs, ignores blank lines and comments.
    Supports quoted values.
    """
    path = Path(dotenv_path)
    if not path.is_file():
        die(f".env file not found: {path}. Create a .env file in this folder (see README).")

    data: Dict[str, str] = {}
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip()

        # remove optional quotes
        if (len(v) >= 2) and ((v[0] == v[-1]) and v[0] in ("'", '"')):
            v = v[1:-1]

        data[k] = v
    return data


def require_cfg(cfg: Dict[str, str], key: str) -> str:
    v = cfg.get(key)
    if not v:
        die(f"Missing required config in .env: {key}")
    return v


def optional_cfg(cfg: Dict[str, str], key: str, default: str) -> str:
    v = cfg.get(key)
    if v is None or v.strip() == "":
        return default
    return v.strip()


def truthy(value: str) -> bool:
    return value.strip().lower() in ("1", "true", "yes", "y", "on")


# ---------------------------------------------------------------------------
# Logging helpers
# ---------------------------------------------------------------------------

def log(msg: str) -> None:
    """
    Print a timestamped log line.

    Input:
        msg: message string

    Output:
        None (prints to stdout)
    """
    now = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"[{now}] {msg}")


def die(msg: str, code: int = 1) -> None:
    """
    Print an error message (timestamped) then exit.

    Input:
        msg: error description
        code: exit code (default: 1)

    Output:
        None (exits the process)
    """
    log(f"ERROR: {msg}")
    sys.exit(code)


def parse_args(tz: ZoneInfo) -> dt.date:
    """
    Parse command-line date argument.

    Supported formats:
        - YYYY-MM-DD     e.g. 2026-01-13
        - DD/MM/YYYY     e.g. 13/01/2026
        - DD/MM/YY       e.g. 13/01/26

    If no argument is provided, defaults to *today* in the user's TRACKER_TIMEZONE.

    Output:
        A dt.date object representing the local target day.
    """
    if len(sys.argv) >= 2 and sys.argv[1].strip():
        s = sys.argv[1].strip()
        fmts = ["%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"]
        for f in fmts:
            try:
                return dt.datetime.strptime(s, f).date()
            except Exception:
                pass
        raise ValueError(f"Unsupported date format: {s}. Use YYYY-MM-DD or DD/MM/YYYY")
    return dt.datetime.now(tz).date()


# ---------------------------------------------------------------------------
# GitHub REST helpers
# ---------------------------------------------------------------------------

def gh_headers(token: str) -> Dict[str, str]:
    """
    Build GitHub request headers.

    Input:
        token: GitHub personal access token (PAT)

    Output:
        dict of HTTP headers used in all GitHub API calls.
    """
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "github-daily-tracker"
    }


def request_get(url: str, headers: Dict[str, str], params=None, retries: int = 3, backoff: float = 2.0) -> requests.Response:
    """
    Perform a GET request with basic retry logic for transient GitHub errors.

    Retries on:
        - 429, 502, 503, 504

    Special handling:
        - If 403 contains "rate limit", exits with a helpful message.

    Input:
        url: endpoint
        headers: HTTP headers dict
        params: query params dict
        retries: max retries
        backoff: seconds multiplier for incremental backoff

    Output:
        requests.Response (already status-checked via raise_for_status)
    """
    for attempt in range(1, retries + 1):
        r = requests.get(url, headers=headers, params=params)
        if r.status_code in (429, 502, 503, 504):
            log(f"HTTP {r.status_code} from GitHub. Retry {attempt}/{retries}...")
            time.sleep(backoff * attempt)
            continue
        if r.status_code == 403 and 'rate limit' in r.text.lower():
            die("GitHub rate limit hit (403). Try later or use a token with higher limits.")
        r.raise_for_status()
        return r
    # If we exhausted retries
    r.raise_for_status()
    return r


def get_all_pages(url: str, headers: Dict[str, str], params=None, max_pages: int = 10) -> List[Dict[str, Any]]:
    """
    Fetch and concatenate paginated GitHub REST API results.

    This follows 'Link' headers for pagination.

    Input:
        url: initial URL
        headers: HTTP headers dict
        params: initial query params
        max_pages: safety cap on total pages to fetch

    Output:
        A list of JSON objects from all fetched pages.
    """
    out: List[Dict[str, Any]] = []
    page = 1
    while url and page <= max_pages:
        r = request_get(url, headers=headers, params=params)
        out.extend(r.json())
        if 'next' in r.links:
            url = r.links['next']['url']
            params = None  # because next page URL already includes params
        else:
            url = None
        page += 1
    return out


# ---------------------------------------------------------------------------
# GitHub metrics: Search API counts
# ---------------------------------------------------------------------------

def search_count(query: str, headers: Dict[str, str], api_base: str) -> int:
    """
    Run a GitHub Search API query and return total_count.

    Input:
        query: search query string, e.g. "repo:org/repo is:pr author:me created:2026-01-13"
        headers: GitHub headers
        api_base: GitHub API base URL (api.github.com or GHES /api/v3)

    Output:
        Integer total_count returned by GitHub.

    Note:
        GitHub Search API caps results at 1000. If total_count exceeds 1000, this
        value is still reported but a warning is logged because pagination can
        only access the first 1000 items.
    """
    url = f"{api_base}/search/issues"
    params = {"q": query, "per_page": 1}
    r = request_get(url, headers=headers, params=params)
    payload = r.json()
    total = int(payload.get("total_count", 0))
    if total > 1000:
        log(f"WARNING: Search query exceeded 1000 results ({total}); some results may be inaccessible. Query: {query}")
    return total


def count_open_counts_asof(owner: str, repo: str, day: dt.date, headers: Dict[str, str], api_base: str, tz: ZoneInfo) -> Tuple[int, int]:
    """
    Compute open issues and open PRs snapshot "as-of end of target day".

    Logic (split into two queries to avoid undocumented boolean grouping):
        Query A: created<=day AND still open now  (lower bound — may undercount if closed later)
        Query B: created<=day AND closed AFTER day (were open on that day but closed since)
        Total = A + B

    NOTE: This is an APPROXIMATION. Items closed before/on `day` and later
    reopened (and currently open) will be incorrectly counted in Query A.
    Exact reconstruction would require event/timeline traversal per item.

    Input:
        owner, repo: target repo
        day: date object
        headers: GitHub headers

    Output:
        (open_issues_count, open_prs_count) as integers
    """
    _, end_utc = day_bounds_utc(day, tz)
    end_iso = fmt_iso_z(end_utc)

    # Issues: still open + created on or before that day
    open_issues_now = search_count(
        f"repo:{owner}/{repo} is:issue is:open created:<{end_iso}", headers, api_base,
    )
    # Issues: closed after that day (were open on that day) + created on or before
    closed_issues_after = search_count(
        f"repo:{owner}/{repo} is:issue is:closed closed:>={end_iso} created:<{end_iso}", headers, api_base,
    )

    # PRs: still open + created on or before that day
    open_prs_now = search_count(
        f"repo:{owner}/{repo} is:pr is:open created:<{end_iso}", headers, api_base,
    )
    # PRs: closed after that day + created on or before
    closed_prs_after = search_count(
        f"repo:{owner}/{repo} is:pr is:closed closed:>={end_iso} created:<{end_iso}", headers, api_base,
    )

    return open_issues_now + closed_issues_after, open_prs_now + closed_prs_after


# ---------------------------------------------------------------------------
# GitHub metrics: Search API for triage/resolved, Commits API for commits
# ---------------------------------------------------------------------------

def count_issues_triaged(owner: str, repo: str, username: str, day: dt.date, headers: Dict[str, str], api_base: str, tz: ZoneInfo, max_pages: int = DEFAULT_MAX_PAGES) -> int:
    """
    Count unique issues (not PRs) in the target repo where the user commented on that day.

    Uses the Issue Comments API (GET /repos/.../issues/comments?since=DAY) and
    filters by created_at within the day + user login. This is more accurate than
    Search API's commenter:USER updated:DAY which only checks the issue's last
    updated_at timestamp — that overcounts when someone else updates an issue the
    user previously commented on, and undercounts for historical dates when
    updated_at has moved past the target day.

    Input:
        owner, repo: target repository
        username: GitHub username
        day: target date
        headers: GitHub headers

    Output:
        Integer count of unique issues commented on.
    """
    day_start, day_end = day_bounds_utc(day, tz)
    since_str = fmt_iso_z(day_start)

    url = f"{api_base}/repos/{owner}/{repo}/issues/comments"
    params: Dict[str, Any] = {"since": since_str, "sort": "created", "direction": "desc", "per_page": 100}

    issue_numbers: set = set()
    page_url: str | None = url
    page_params = params

    pages_seen = 0
    reached_day_start = False
    for _ in range(max_pages):
        pages_seen += 1
        r = request_get(page_url, headers, params=page_params)
        comments = r.json()
        if not comments:
            reached_day_start = True
            break

        done = False
        for c in comments:
            created = dt.datetime.strptime(c["created_at"], "%Y-%m-%dT%H:%M:%SZ")
            if created >= day_end:
                continue  # newer than target day, skip
            if created < day_start:
                done = True  # older than target day — all subsequent are older too
                break

            if c.get("user", {}).get("login") != username:
                continue
            # Skip PR comments (html_url contains /pull/ for PRs, /issues/ for issues)
            if "/pull/" in c.get("html_url", ""):
                continue

            issue_url = c.get("issue_url", "")
            try:
                issue_numbers.add(int(issue_url.rsplit("/", 1)[-1]))
            except (ValueError, IndexError):
                pass

        if done:
            reached_day_start = True
            break
        if "next" in r.links:
            page_url = r.links["next"]["url"]
            page_params = None
        else:
            reached_day_start = True
            break

    if not reached_day_start:
        log(
            f"WARNING: count_issues_triaged hit max_pages={max_pages} ({pages_seen} pages fetched) "
            f"without reaching day_start={day_start.isoformat()}. Result may be undercounted. "
            f"Increase TRACKER_MAX_PAGES if needed."
        )

    return len(issue_numbers)


def count_issues_resolved(owner: str, repo: str, username: str, day: dt.date, headers: Dict[str, str], api_base: str, tz: ZoneInfo, max_pages: int = DEFAULT_MAX_PAGES) -> int:
    """
    Count issues the user closed on that day in the target repo.

    Strategy:
        1. Search API finds all issues closed on that day in the repo.
        2. For each, fetch issue events to verify the user was the closer.

    This is more accurate than involves:USER which matches anyone who
    authored, commented, was assigned, or was mentioned — not just the closer.

    Input:
        owner, repo: target repository
        username: GitHub username
        day: target date
        headers: GitHub headers

    Output:
        Integer count of issues the user closed.
    """
    day_start, day_end = day_bounds_utc(day, tz)
    start_iso = fmt_iso_z(day_start)
    end_inclusive_iso = fmt_iso_z(day_end - dt.timedelta(seconds=1))

    # Step 1: Find all issues closed within the local day window (paginated, capped by Search API at 1000)
    query = f"repo:{owner}/{repo} is:issue is:closed closed:{start_iso}..{end_inclusive_iso}"
    search_url: str | None = f"{api_base}/search/issues"
    search_params: Dict[str, Any] = {"q": query, "per_page": 100}
    items: List[Dict[str, Any]] = []
    pages = 0
    while search_url and pages < max_pages:
        r = request_get(search_url, headers, params=search_params)
        payload = r.json()
        items.extend(payload.get("items", []))
        if pages == 0 and int(payload.get("total_count", 0)) > 1000:
            log(
                f"WARNING: >1000 issues closed in window {start_iso}..{end_inclusive_iso}; Search API can only return the first 1000."
            )
        if "next" in r.links:
            search_url = r.links["next"]["url"]
            search_params = None
        else:
            search_url = None
        pages += 1

    # Step 2: For each closed issue, check who closed it (paginate events)
    count = 0
    for issue in items:
        number = issue["number"]
        events_url: str | None = f"{api_base}/repos/{owner}/{repo}/issues/{number}/events"
        events_params: Dict[str, Any] = {"per_page": 100}
        events: List[Dict[str, Any]] = []
        ev_pages = 0
        while events_url and ev_pages < max_pages:
            ev_r = request_get(events_url, headers, params=events_params)
            events.extend(ev_r.json())
            if "next" in ev_r.links:
                events_url = ev_r.links["next"]["url"]
                events_params = None
            else:
                events_url = None
            ev_pages += 1

        # Walk events in reverse to find the most recent "closed" event on that day
        for ev in reversed(events):
            if ev.get("event") != "closed":
                continue
            ev_time = dt.datetime.strptime(ev["created_at"], "%Y-%m-%dT%H:%M:%SZ")
            if day_start <= ev_time < day_end:
                if ev.get("actor", {}).get("login") == username:
                    count += 1
            break  # only check the most recent close event

    return count


def count_commits(owner: str, repo: str, username: str, day: dt.date, headers: Dict[str, str], api_base: str, tz: ZoneInfo, max_pages: int = DEFAULT_MAX_PAGES) -> int:
    """
    Count commits authored by the user in the target repo on the given local day.

    Uses the Commits API: GET /repos/{owner}/{repo}/commits?author={user}&since=&until=
    This avoids all PushEvent/compare issues (force pushes, new branches, 90-day limit).
    The day window is computed in the user's TRACKER_TIMEZONE.

    Output:
        Integer count of commits.
    """
    start_utc, end_utc = day_bounds_utc(day, tz)
    since = fmt_iso_z(start_utc)
    until = fmt_iso_z(end_utc)

    commits_url = f"{api_base}/repos/{owner}/{repo}/commits"
    params = {
        "author": username,
        "since": since,
        "until": until,
        "per_page": 100,
    }
    commits = get_all_pages(commits_url, headers, params=params, max_pages=max_pages)
    return len(commits)


# ---------------------------------------------------------------------------
# Aggregate daily metrics
# ---------------------------------------------------------------------------

def count_metrics(
    owner: str,
    repo: str,
    username: str,
    day: dt.date,
    headers: Dict[str, str],
    tz: ZoneInfo,
    api_base: str = DEFAULT_API,
    include_open_snapshot: bool = True,
    max_pages: int = DEFAULT_MAX_PAGES,
) -> Dict[str, int]:
    """
    Gather all metrics needed for one date row.

    All queries are scoped to the target repo via Search API or Commits API.
    The day window is computed in the user's TRACKER_TIMEZONE so contributions
    near local midnight land on the expected tracker date.

    Output:
        dict:
            issues_triaged, issues_resolved, prs_created, prs_merged,
            commits, open_issues, open_prs
    """
    start_utc, end_utc = day_bounds_utc(day, tz)
    start_iso = fmt_iso_z(start_utc)
    end_inclusive_iso = fmt_iso_z(end_utc - dt.timedelta(seconds=1))
    window = f"{start_iso}..{end_inclusive_iso}"

    # 1) PR counts via Search API
    log(f"Fetching PR counts (created / merged) via search [{window}]...")
    prs_created = search_count(f"repo:{owner}/{repo} is:pr author:{username} created:{window}", headers, api_base)
    prs_merged = search_count(f"repo:{owner}/{repo} is:pr author:{username} merged:{window}", headers, api_base)

    # 2) Issues triaged (commented on) via Issue Comments API
    log("Fetching issues triaged (commented on) via search...")
    issues_triaged = count_issues_triaged(owner, repo, username, day, headers, api_base, tz, max_pages)

    # 3) Issues resolved (closed) via Search API
    log("Fetching issues resolved (closed) via search...")
    issues_resolved = count_issues_resolved(owner, repo, username, day, headers, api_base, tz, max_pages)

    # 4) Commits via Commits API
    log("Fetching commits count via Commits API...")
    commits = count_commits(owner, repo, username, day, headers, api_base, tz, max_pages)

    metrics: Dict[str, int] = {
        "issues_triaged": issues_triaged,
        "issues_resolved": issues_resolved,
        "prs_created": prs_created,
        "prs_merged": prs_merged,
        "commits": commits,
    }

    # 5) Open snapshot as-of that day (optional — 4 extra search calls)
    if include_open_snapshot:
        log("Fetching open issues/PRs counts as-of target day...")
        open_issues, open_prs = count_open_counts_asof(owner, repo, day, headers, api_base, tz)
        metrics["open_issues"] = open_issues
        metrics["open_prs"] = open_prs
    else:
        log("Skipping open issues/PRs snapshot (TRACKER_INCLUDE_OPEN_SNAPSHOT=false).")

    return metrics


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def find_or_create_row(ws, day: dt.date, date_format: str = DEFAULT_DATE_FORMAT) -> int:
    """
    Find an existing row whose Date column equals 'day', otherwise append a new row.

    Side effect: normalizes the Excel number format on EVERY Date cell in the
    sheet so the whole column stays consistent (fixes drift when Excel
    occasionally re-applies a locale-default format like 'mm-dd-yy' to recently
    written rows).

    Assumes:
        Column A is Date
        Row 1 is header
        Data starts from row 2

    Input:
        ws: openpyxl worksheet
        day: target date
        date_format: Excel number format for the Date cell (e.g., "dd/mm/yyyy")

    Output:
        row index (int) where data should be written.
    """
    match_row = None
    last_data_row = 1  # header
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, 1)
        v = cell.value
        if isinstance(v, dt.datetime):
            v = v.date()
        if isinstance(v, dt.date):
            # Force-normalize format on every existing Date cell
            cell.number_format = date_format
            if v == day and match_row is None:
                match_row = r
            last_data_row = r

    if match_row is not None:
        return match_row

    r = last_data_row + 1
    ws.cell(r, 1).value = day
    ws.cell(r, 1).number_format = date_format
    return r


def validate_xlsx(path: str) -> None:
    """
    Validate workbook path before openpyxl loads it.

    Checks:
        - exists
        - is file (not directory)
        - reasonable size (> 1000 bytes)
        - is a ZIP container (xlsx is a zip)

    Input:
        path: file path

    Output:
        None (dies on failure)
    """
    if not os.path.exists(path):
        die(f"Workbook not found: {path}. Put the .xlsx in this folder or set TRACKER_OUT in .env.")
    if os.path.isdir(path):
        die(f"Workbook path is a directory, not a file: {path}")
    size = os.path.getsize(path)
    if size < 1000:
        die(f"Workbook file looks too small ({size} bytes): {path}. Re-download the .xlsx.")
    if not zipfile.is_zipfile(path):
        die("Workbook is not a valid .xlsx (zip) file.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    """
    Script entry point.

    Steps:
        1) parse date
        2) validate token + workbook
        3) read Config sheet (owner/repo/username)
        4) compute metrics from GitHub
        5) write to the correct date row
        6) save workbook

    Output:
        None (prints logs, writes Excel)
    """
    script_dir = Path(__file__).resolve().parent
    env_cfg = load_dotenv(str(script_dir / ".env"))

    token = require_cfg(env_cfg, "GITHUB_TOKEN")
    xlsx = require_cfg(env_cfg, "TRACKER_OUT")
    owner = require_cfg(env_cfg, "GITHUB_OWNER")
    repo = require_cfg(env_cfg, "GITHUB_REPO")
    username = require_cfg(env_cfg, "GITHUB_USERNAME")
    sheet_name = require_cfg(env_cfg, "TRACKER_SHEET")

    # Optional, customizable settings
    api_base = optional_cfg(env_cfg, "GITHUB_API_BASE_URL", DEFAULT_API).rstrip("/")
    date_format = optional_cfg(env_cfg, "TRACKER_DATE_FORMAT", DEFAULT_DATE_FORMAT)
    include_open_snapshot = truthy(optional_cfg(env_cfg, "TRACKER_INCLUDE_OPEN_SNAPSHOT", "true"))
    tz_name = optional_cfg(env_cfg, "TRACKER_TIMEZONE", DEFAULT_TIMEZONE)
    tz = resolve_tz(tz_name)
    try:
        max_pages = int(optional_cfg(env_cfg, "TRACKER_MAX_PAGES", str(DEFAULT_MAX_PAGES)))
        if max_pages < 1:
            raise ValueError
    except ValueError:
        die("TRACKER_MAX_PAGES must be a positive integer.")

    # parse the date arg using the resolved timezone for the "today" default
    try:
        day = parse_args(tz)
    except Exception as e:
        die(str(e))

    # Resolve relative TRACKER_OUT against the script directory so the tool
    # works regardless of the current working directory.
    xlsx_path = Path(xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = script_dir / xlsx_path
    xlsx = str(xlsx_path)

    log(f"Using workbook: {os.path.abspath(xlsx)}")
    validate_xlsx(xlsx)

    log("Opening workbook...")
    try:
        wb = load_workbook(xlsx)
    except zipfile.BadZipFile:
        die(f"BadZipFile: {xlsx} is not a real .xlsx. Re-download the workbook and try again.")

    if sheet_name not in wb.sheetnames:
        die(f"Worksheet '{sheet_name}' not found in workbook. Did you run init_tracker.py?")

    # Optional: enforce init_tracker-created Config sheet exists
    if "Config" not in wb.sheetnames:
        die("Sheet 'Config' not found in workbook. Did you run init_tracker.py?")

    log(f"Target date: {day:%Y-%m-%d} | TZ: {tz_name} | Repo: {owner}/{repo} | User: {username} | Sheet: {sheet_name} | API: {api_base}")

    headers = gh_headers(token)
    metrics = count_metrics(
        owner, repo, username, day, headers, tz,
        api_base=api_base,
        include_open_snapshot=include_open_snapshot,
        max_pages=max_pages,
    )

    ws = wb[sheet_name]
    row = find_or_create_row(ws, day, date_format)

    log(f"Writing metrics into row {row}...")
    # A Date | B Issues Triaged | C Issues Resolved | D PRs Created
    # E PRs Merged | F Commits  | G Open Issues     | H Open PRs
    # (I Release, J Notes are manual — never touched)
    ws.cell(row, 2).value = metrics["issues_triaged"]
    ws.cell(row, 3).value = metrics["issues_resolved"]
    ws.cell(row, 4).value = metrics["prs_created"]
    ws.cell(row, 5).value = metrics["prs_merged"]
    ws.cell(row, 6).value = metrics["commits"]
    if "open_issues" in metrics:
        ws.cell(row, 7).value = metrics["open_issues"]
    if "open_prs" in metrics:
        ws.cell(row, 8).value = metrics["open_prs"]

    # Update "Last Updated (UTC)" in Config
    cfg_ws = wb["Config"]
    cfg_ws["B5"].value = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")

    log("Saving workbook...")
    wb.save(xlsx)

    log(f"DONE. Updated {os.path.basename(xlsx)} for {day.strftime('%Y-%m-%d')}: {metrics}")


if __name__ == "__main__":
    main()
